from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import random
from datetime import timedelta
from django.utils import timezone
from ..forms import UserRegistrationForm, EmailVerificationForm
from ..models import User
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime


# Create your views here.

def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            user.verification_code_created_at = timezone.now()
            user.save()
            
            # Отправка кода на email
            send_mail(
                'Подтверждение регистрации',
                f'Ваш код подтверждения: {user.verification_code}',
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            
            messages.success(request, 'Код подтверждения отправлен на ваш email')
            return redirect('verify_email', user_id=user.id)
    else:
        form = UserRegistrationForm()
    return render(request, 'users/register.html', {'form': form})

def verify_email(request, user_id):
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        messages.error(request, 'Пользователь не найден')
        return redirect('register')
        
    if request.method == 'POST':
        form = EmailVerificationForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data['verification_code']
            if (user.verification_code == code and 
                user.verification_code_created_at + timedelta(minutes=30) > timezone.now()):
                user.is_active = True
                user.is_email_verified = True
                user.verification_code = None
                user.verification_code_created_at = None
                user.save()
                login(request, user)
                messages.success(request, 'Email успешно подтвержден!')
                return redirect('user_list')
            else:
                messages.error(request, 'Неверный код или код устарел')
    else:
        form = EmailVerificationForm()
    return render(request, 'users/verify_email.html', {'form': form, 'user': user})

@login_required
def user_list(request):
    if not request.user.is_email_verified:
        messages.error(request, 'Для доступа к списку пользователей необходимо подтвердить email')
        return redirect('home')
    
    # Получаем параметры фильтрации из GET-запроса
    role_filter = request.GET.get('role', '')
    search_query = request.GET.get('search', '')
    
    # Базовый queryset
    users = User.objects.all()
    
    #  фильтры
    if role_filter:
        users = users.filter(role=role_filter)
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(username__icontains=search_query)
        )
    
    # Сортировка
    users = users.order_by('-date_joined')
    
    # Пагинация
    paginator = Paginator(users, 10)  # 10 пользователей на страницу
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Контекст для шаблона
    context = {
        'page_obj': page_obj,
        'total_users': users.count(),
        'current_role': role_filter,
        'search_query': search_query,
    }
    
    return render(request, 'users/user_list.html', context)

@login_required
def export_users_excel(request):
    # Получаем те же параметры фильтрации, что и в списке
    role_filter = request.GET.get('role')
    search_query = request.GET.get('search')
    
    # Базовый queryset
    users = User.objects.all()
    
    # Применяем фильтры
    if role_filter:
        users = users.filter(role=role_filter)
    if search_query:
        users = users.filter(
            Q(email__icontains=search_query) |
            Q(username__icontains=search_query)
        )
    
    # Создаем новый Excel-файл https://openpyxl.readthedocs.io/en/stable/
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Стили для заголовков
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # Заголовки
    headers = [
        'ID', 'Email', 'Роль', 'Дата регистрации', 'Статус верификации',
        'Посещаемость (%)', 'Количество уроков', 'Общий балл'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Данные
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.id)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=user.get_role_display())
        ws.cell(row=row, column=4, value=user.date_joined.strftime("%d.%m.%Y %H:%M"))
        ws.cell(row=row, column=5, value="Подтвержден" if user.is_email_verified else "Не подтвержден")
        
        if user.role == User.Role.STUDENT:
            ws.cell(row=row, column=6, value=user.attendance_percentage)
            ws.cell(row=row, column=7, value=user.lessons_count)
            ws.cell(row=row, column=8, value=user.total_score)
        else:
            ws.cell(row=row, column=6, value="-")
            ws.cell(row=row, column=7, value="-")
            ws.cell(row=row, column=8, value="-")
    
    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Создаем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Сохраняем файл
    wb.save(response)
    return response